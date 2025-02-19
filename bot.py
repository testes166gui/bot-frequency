import telebot # Usado para criar um bot do Telegram.   
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton # Permite criar botões no bot do Telegram.
from datetime import datetime, timedelta # Trabalham com datas e horários.
import pandas as pd # Manipula dados em formato de tabela (como planilhas do Excel).
from fpdf import FPDF # Cria arquivos PDF.
import threading # Permite executar múltiplas tarefas ao mesmo tempo.
from openpyxl import load_workbook # Trabalha com arquivos do Excel (.xlsx).
from dotenv import load_dotenv # Carrega variáveis de ambiente (como senhas armazenadas separadamente).
import time
import schedule
import sys
from PyPDF2 import PdfMerger

def agendar_tarefas():
    while True:
        schedule.run_pending()
        time.sleep(2)

def main():
    while True:

        import os
        try:
            # Diretório onde os arquivos individuais estão armazenados
            PASTA_ARQUIVOS = "arquivos_individuais"
            load_dotenv()  # Carrega as variáveis de ambiente do arquivo .env
            print("Variáveis de ambiente carregadas.")

            CHAVE_API = os.getenv("TELEGRAM_API_KEY") # Obtém a chave do bot do Telegram do arquivo .env.
            bot = telebot.TeleBot(CHAVE_API) # Cria um bot do Telegram usando essa chave.
            print("Bot do Telegram inicializado.")

            # Lista de IDs dos gerentes que devem receber a mensagem automática às 08:00
            '''gerentes_ids = {1243989891,6656539611, 7985041438, 7570096590, 691282313, 7738937101, 8021490944, 7936163663, 1289492894, 633515141, 8197116783, 7903571530, 8104484593, 682569652, 7937663470, 7970426157, 6694394827, 754823943 }  # Substitua pelos IDs reais dos gerentes

            # Função para enviar mensagem inicial para os gerentes
            def iniciar_conversa_para_gerentes():
                for gerente_id in gerentes_ids:
                    try:
                        bot.send_message(gerente_id, "Bom dia! Por favor, inicie a marcação de presença dos funcionários.", reply_markup=menu_escolha_loja())
                    except Exception as e:
                        print(f"Erro ao enviar mensagem para {gerente_id}: {e}")

            # Agendar envio diário às 08:00
            schedule.every().day.at("08:00").do(iniciar_conversa_para_gerentes)'''

            usuarios = {} # Cria um dicionário (estrutura de dados) vazio para armazenar informações de usuários interagindo com o bot.

            # Dicionário global para armazenar as informações de todas as lojas ao longo do dia
            informacoes_diarias = {}


            # Iniciar o agendamento em uma thread separada
            tarefa_agendamento = threading.Thread(target=agendar_tarefas)
            tarefa_agendamento.daemon = True  # Isso garante que a thread seja fechada quando o programa principal for finalizado
            tarefa_agendamento.start()
            # Lista de lojas autorizadas (número da loja: nome e funcionários)
            lojasCadastradas = {
                "1": {"nome": "Loja 1", "Funcionários": ["Lilian", "Tallyta", "Samara", "Regina", "Daiany"]},
                "4": {"nome": "Loja 4", "Funcionários": ["Joelma Mendes", "Antônio", "Elidalva", "Mikaele", "Carlos"]},
                "5": {"nome": "Loja 5", "Funcionários": ["Jonathan", "Julio Rocha", "Samuel", "Herbert", "Wellington", "Natalia", "Cleidiane", "Nayana", "Eduardo"]},
            }
            # Função para criar o menu principal
            def menu_principal(): # Cria um menu interativo com um botão chamado "Finalizar".
                markup = InlineKeyboardMarkup() # Cria um teclado inline (botões dentro do chat).
                markup.add(InlineKeyboardButton("Finalizar", callback_data="finalizar_conversa")) # Cria um botão que, ao ser clicado, envia a informação "finalizar_conversa" ao bot.
                return markup # Retorna o teclado criado
            
            # Função para garantir que o diretório exista
            def garantir_diretorio(diretorio):
                if not os.path.exists(diretorio):
                    os.makedirs(diretorio)

            # Função para gerar relatórios em Excel
            def gerar_relatorio_excel(usuario_id):
                garantir_diretorio("arquivos_individuais")  # Certifique-se de que o diretório existe
                data = []
                data_atual = datetime.now().strftime("%H:%M %d/%m/%Y")
                nome_loja = lojasCadastradas[usuarios[usuario_id]["loja"]]["nome"]

                if not usuarios[usuario_id]["ausencias"]:
                    data.append({
                        "Loja": nome_loja,
                        "Data": data_atual,
                        "Mensagem": "Todos entraram no horário normal"
                    })
                else:
                    for funcionario, dados in usuarios[usuario_id]["ausencias"].items():
                        # Adicionando dados ao Excel...
                        row = { 
                            "Loja": nome_loja, 
                            "Data": data_atual, 
                            "Funcionário": funcionario,
                            "Motivo": dados.get('motivo', None),
                            "Motivo do Atraso": dados.get('motivo_atraso', None),
                            "Tempo de Atraso (horas)": dados.get('tempo_atraso', {}).get('tempo_atraso_horas', None),
                            "Tempo de Atraso (min)": dados.get('tempo_atraso', {}).get('tempo_atraso_minutos', None),
                            "Tempo que saiu mais cedo (horas)": dados.get('tempo_saiu_mais_cedo', {}).get('tempo_saiu_horas', None),
                            "Tempo que saiu mais cedo (min)": dados.get('tempo_saiu_mais_cedo', {}).get('tempo_saiu_minutos', None),
                            "Motivo de saída antecipada": dados.get('justificativa_cedo', None),
                            "Atestado": "Enviado" if dados.get('atestado', False) else None,
                            "Prazo Atestado": "Prazo de 48 horas para o atestado ser entregue" if dados.get('prazo_atestado', False) else None,  # Corrigido nome da chave
                            "Falta Acordada": "Sim" if dados.get('acordado', False) else None,  # Ajuste correto
                            "Justificativa": dados.get('justificativa', None),
                            
                            }
                        data.append(row)

                df = pd.DataFrame(data).dropna(axis=1, how='all')
                
                
                # Ajustar a largura das colunas automaticamente
                nome_arquivo = f"arquivos_individuais/relatorio_{nome_loja.replace(' ', '')}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}_{usuario_id}.xlsx"
                try:
                    df.to_excel(nome_arquivo, index=False)
                    print(f"Relatório Excel salvo em {nome_arquivo}")
                except Exception as e:
                    print(f"Erro ao salvar o Excel: {e}")
                ajustar_largura_colunas(nome_arquivo)

                

            # Ajusta a largura das colunas do Excel baseado no conteúdo
            def ajustar_largura_colunas(filename):
                if not os.path.exists(filename):
                    print(f"Arquivo {filename} não encontrado.")
                    return
                
                try:
                    wb = load_workbook(filename)
                    # Restante do código...
                except Exception as e:
                    print(f"Erro ao carregar o arquivo: {e}")
                
                wb = load_workbook(filename)
                ws = wb.active
                #  Abre o arquivo Excel e seleciona a planilha ativa.

                for col in ws.columns: # Percorre todas as colunas da planilha.
                    max_length = 0 #  Inicia uma variável max_length para medir o tamanho do maior texto da coluna.
                    col_letter = col[0].column_letter #obtém a letra da coluna

                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                            # Percorre todas as células da coluna e mede o tamanho do maior conteúdo
                    
                    ws.column_dimensions[col_letter].width = max_length + 2 # Ajusta a largura da coluna para caber o maior texto mais um pequeno espaço extra.

                wb.save(filename)

            import os

            # Função para criar o menu de motivo de ausência com botões inline
            def menu_motivo_ausencia(funcionario):
                markup = InlineKeyboardMarkup(row_width=2) # Cria um conjunto de botões inline (dentro do chat), com 2 colunas.
                motivos = ["Atraso", "Atestado", "Faltou", "Folga", "Saiu mais cedo", "Fora do horário"] # Lista os motivos disponíveis.

                for motivo in motivos:
                    # O callback data inclui o nome do funcionário e o motivo
                    markup.add(InlineKeyboardButton(motivo, callback_data=f"motivo_{funcionario}_{motivo}"))
                    # Cria um botão para cada motivo.

                return markup

            # Função para criar o menu de escolha de loja
            def menu_escolha_loja():
                markup = InlineKeyboardMarkup(row_width=2)
                for numero, loja in lojasCadastradas.items():
                    markup.add(InlineKeyboardButton(loja["nome"], callback_data=f"loja_{numero}"))
                return markup

            # Função para criar a lista de funcionários com checkboxes
            def menu_funcionarios(usuario_id, loja):
                markup = InlineKeyboardMarkup(row_width=1) # InlineKeyboardMarkup(row_width=1) define que cada linha do menu terá apenas 1 botão.
                funcionarios = loja["Funcionários"] # Pega a lista de funcionários da loja selecionada.
                
                # Verificar se o usuário tem a chave "funcionarios", caso contrário, inicializar
                if "funcionarios" not in usuarios[usuario_id]:
                    usuarios[usuario_id]["funcionarios"] = {}
                
                # Inicializando as informações dos funcionários para o usuário se ainda não foi feito
                for funcionario in funcionarios:
                    if funcionario not in usuarios[usuario_id]["funcionarios"]:
                        usuarios[usuario_id]["funcionarios"][funcionario] = {"presente": False}
                        # Se o funcionário ainda não está no registro do usuário, adiciona ele com a presença marcada como "FALSO" (False).
                        # Isso significa que, por padrão, todos os funcionários começam como "ausentes".
                    
                    # Definindo o texto do botão (com ou sem ✔️)
                    if usuarios[usuario_id]["funcionarios"][funcionario]["presente"]:
                        button_text = f"✅ {funcionario}"
                    else:
                        button_text = f"{funcionario}"
                    
                    markup.add(InlineKeyboardButton(button_text, callback_data=f"presente_{funcionario}"))
                
                # Botão para enviar a presença
                markup.add(InlineKeyboardButton("✅ Enviar Presença", callback_data="enviar_presenca"))
                
                return markup

            # Handler para iniciar a conversa e perguntar pela loja
            @bot.message_handler(func=lambda mensagem: "loja" not in usuarios.get(mensagem.chat.id, {})) 
            # "loja" not in ...: verifica se o usuário ainda não escolheu uma loja, usuarios.get(mensagem.chat.id, {}): pega os dados do usuário ou um dicionário vazio ({}) se ele ainda não tiver interagido.
            def iniciar_conversa(mensagem):
                bot.send_message(mensagem.chat.id, "Selecione o número da sua loja:", reply_markup=menu_escolha_loja()) # Usa a função menu_escolha_loja() para gerar os botões.

            # Handler para registrar a loja escolhida
            @bot.callback_query_handler(func=lambda call: call.data.startswith('loja_'))
            def registrar_loja(call):
                numero_loja = call.data.split("_")[1]
                # Extrai o número da loja do callback, Se o callback for "loja_2", numero_loja será "2".

                if numero_loja in lojasCadastradas: #  Verifica se a loja realmente existe.
                    loja = lojasCadastradas[numero_loja] # Obtém os dados da loja escolhida.
                    usuario_id = call.message.chat.id # Pega o ID do usuário que fez a seleção.
                    
                    # Inicializando o dicionário para o usuário, caso ainda não tenha sido inicializado
                    if usuario_id not in usuarios: # Se o usuário ainda não estiver registrado, cria um novo registro para ele.
                        usuarios[usuario_id] = {"loja": numero_loja, "nome_funcionario": None, "ausencias": {}, "funcionarios": {}}
                    else:
                        # Se já existe o usuário, reinicia as informações de funcionários e presença
                        usuarios[usuario_id]["funcionarios"] = {}  # Limpa os funcionários antigos
                        usuarios[usuario_id]["ausencias"] = {}     # Limpa as ausências antigas
                        usuarios[usuario_id]["loja"] = numero_loja  # Atualiza para a nova loja

                    # Criar a lista de funcionários com caixas de seleção
                    markup = menu_funcionarios(call.message.chat.id, loja)

                    # Confirmar a loja registrada e enviar a lista de funcionários
                    bot.send_message(call.message.chat.id, f"Loja registrada: {loja['nome']}. Marque os funcionários que entraram no seu horário normal de trabalho:\nNão digite nada nessa parte!", reply_markup=markup)
                else:
                    bot.send_message(call.message.chat.id, "Loja não encontrada! Tente novamente.")

            # Handler para marcar/desmarcar presença de funcionários
            @bot.callback_query_handler(func=lambda call: call.data.startswith('presente_')) # O @bot.callback_query_handler serve para dizer ao bot: "Ei, fique atento! Se alguém clicar em um botão com 'presente_', chame essa função."
            def marcar_desmarcar_presenca(call):
                funcionario = call.data.split("_")[1] #  call.data contém um texto como "presente_João", "presente_Maria", etc.
                # O .split("_")[1] separa o texto em partes e pega a segunda, que é o nome do funcionário.
                # Exemplo: "presente_João" → separa em ["presente", "João"] → pegamos "João".

                usuario_id = call.message.chat.id

                # Verifica se o usuário já tem informações registradas (loja e funcionários)
                if usuario_id not in usuarios:
                    bot.answer_callback_query(call.id, text="Você precisa se registrar primeiro.")
                    return

                # Verifica se a loja foi registrada
                if "loja" not in usuarios[usuario_id] or not usuarios[usuario_id]["loja"]:
                    bot.answer_callback_query(call.id, text="Loja não registrada. Por favor, selecione a loja novamente.")
                    return

                # Inicializar o dicionário de funcionários se necessário
                if "funcionarios" not in usuarios[usuario_id]:
                    usuarios[usuario_id]["funcionarios"] = {}

                # Verifica se o funcionário está no dicionário de presença do usuário
                if funcionario not in usuarios[usuario_id]["funcionarios"]:
                    usuarios[usuario_id]["funcionarios"][funcionario] = {"presente": False}

                # Alternar o estado de presença
                usuarios[usuario_id]["funcionarios"][funcionario]["presente"] = not usuarios[usuario_id]["funcionarios"][funcionario]["presente"]

                # Atualizando a presença do funcionário
                loja = lojasCadastradas[usuarios[usuario_id]["loja"]]
                markup_novo = menu_funcionarios(usuario_id, loja)
                markup_atual = call.message.reply_markup  # Teclado atual da mensagem


                # Verifica se o novo teclado é diferente do atual
                if markup_atual and markup_novo.to_json() == markup_atual.to_json():
                    bot.answer_callback_query(call.id, text="Nenhuma alteração feita.")  # Mensagem de feedback
                    return
                
                # Atualizando a presença visualmente
                bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=markup_novo)
                bot.answer_callback_query(call.id, text=f"{funcionario} presença alternada.")

            # Handler para o envio da presença
            @bot.callback_query_handler(func=lambda call: call.data == "enviar_presenca")
            def enviar_presenca(call):
                usuario_id = call.message.chat.id
                loja = usuarios[usuario_id]["loja"]
                funcionarios = lojasCadastradas[loja]["Funcionários"]

                ausentes = []
                
                # Verificar quais funcionários faltaram
                for funcionario in funcionarios:
                    if not usuarios[usuario_id]["funcionarios"][funcionario]["presente"]:
                        ausentes.append(funcionario)
                
                # Se houver ausentes, perguntar o motivo
                if ausentes:
                    for funcionario in ausentes:
                        bot.send_message(
                            call.message.chat.id,
                            f"{funcionario} não entrou no horário normal de trabalho. Qual o motivo?",
                            reply_markup=menu_motivo_ausencia(funcionario)
                        )
                
                else:
                    bot.send_message(call.message.chat.id, "Todos entraram no horário normal.",
                            reply_markup=menu_principal())
                
            # Handler para selecionar o motivo da ausência
            @bot.callback_query_handler(func=lambda call: call.data.startswith("motivo_"))
            def motivo_ausencia(call):
                usuario_id = call.message.chat.id
                _, funcionario, motivo = call.data.split("_")  # Extrair o funcionário e o motivo do callback data

                #Registrar motivo inicial no dicionário do usuário
                if "ausencias" not in usuarios [usuario_id]:
                    usuarios[usuario_id]["ausencias"] = {}

                usuarios [usuario_id]["ausencias"][funcionario] = {"motivo":motivo}

                #Processar o motivo com base na escolha
                match motivo:
                    case "Atraso":
                        markup = InlineKeyboardMarkup(row_width=1)
                        markup.add(
                            InlineKeyboardButton("Horas", callback_data=f"atraso_horas_{funcionario}"),
                            InlineKeyboardButton("Minutos", callback_data=f"atraso_minutos_{funcionario}")
                        )
                        bot.send_message(
                            usuario_id,
                            f"Quanto tempo de atraso para {funcionario}?",
                            reply_markup=markup
                        )
                        
                    case "Atestado":
                        markup = InlineKeyboardMarkup(row_width=2)
                        markup.add(
                            InlineKeyboardButton("Agora", callback_data=f"atestado_agora_{funcionario}"),
                            InlineKeyboardButton("Depois", callback_data=f"atestado_depois_{funcionario}"),
                            InlineKeyboardButton("Já Enviei", callback_data=f"atestado_ja_{funcionario}")
                        )
                        bot.send_message(
                            usuario_id,
                            f"Deseja enviar o atestado agora ou depois?",
                            reply_markup=markup
                        )
                        
                    case "Faltou":
                        markup = InlineKeyboardMarkup(row_width=2)
                        markup.add(
                            InlineKeyboardButton("Sim", callback_data=f"ausente_sim_{funcionario}"),
                            InlineKeyboardButton("Não", callback_data=f"ausente_nao_{funcionario}")
                        )
                        bot.send_message(
                            usuario_id,
                            f"A falta de {funcionario} foi acordada previamente?",
                            reply_markup=markup
                        )
                    case "Folga":
                        bot.send_message(
                            usuario_id,
                            f"A ausência de {funcionario} foi justificada como folga.",
                            reply_markup=menu_principal()
                        )
                    case "Saiu mais cedo":
                        markup = InlineKeyboardMarkup(row_width=1)
                        markup.add(
                            InlineKeyboardButton("Horas", callback_data=f"cedo_horas_{funcionario}"),
                            InlineKeyboardButton("Minutos", callback_data=f"cedo_minutos_{funcionario}")
                        )
                        bot.send_message(
                            usuario_id,
                            f"Quantos minutos {funcionario} saiu mais cedo?",
                            reply_markup=markup
                        )
                        
                    case "Fora do horário":
                        bot.send_message(
                            usuario_id,
                            f"A ausência de {funcionario} foi justificada por estar fora de horário de trabalho",
                            reply_markup=menu_principal()
                        )


            # Função para registrar tempo de atraso
            @bot.callback_query_handler(func=lambda call: call.data.startswith("atraso_"))
            def registrar_tempo_atraso(call):
                usuario_id = call.message.chat.id
                _, tipo, funcionario = call.data.split("_")

                if tipo == "minutos":
                    markup = InlineKeyboardMarkup(row_width=1)
                    markup.add(
                        InlineKeyboardButton("1 a 15 minutos", callback_data=f"atrasotempo_minutos_1_15_{funcionario}"),
                        InlineKeyboardButton("15 a 45 minutos", callback_data=f"atrasotempo_minutos_15_45_{funcionario}"),
                        InlineKeyboardButton("30 a 59 minutos", callback_data=f"atrasotempo_minutos_30_59_{funcionario}")
                    )
                    bot.send_message(
                        usuario_id,
                        f"Selecione o tempo de atraso para {funcionario}:",
                        reply_markup=markup
                    )

                elif tipo == "horas":
                    markup = InlineKeyboardMarkup(row_width=1)
                    markup.add(
                        InlineKeyboardButton("1 hora", callback_data=f"atrasotempo_horas_1_{funcionario}"),
                        InlineKeyboardButton("2 horas", callback_data=f"atrasotempo_horas_2_{funcionario}"),
                        InlineKeyboardButton("3 horas ou mais", callback_data=f"atrasotempo_horas_3_{funcionario}")
                    )
                    bot.send_message(
                        usuario_id,
                        f"Selecione o tempo de atraso em horas para {funcionario}:",
                        reply_markup=markup
                    )
    
            # Registrar tempo de atraso selecionado pelo usuário
            @bot.callback_query_handler(func=lambda call: call.data.startswith("atrasotempo_"))
            def salvar_tempo_atraso(call):
                usuario_id = call.message.chat.id
                try:
                    # Dividir o callback_data corretamente
                    partes = call.data.split("_")
                    print(f"Partes divididas: {partes}")  # Para depuração

                    if len(partes) < 4:
                        raise ValueError("Formato inválido para callback_data.")

                    unidade = partes[1]  # Unidade de tempo (horas ou minutos)
                    tempo = None
                    funcionario = partes[-1]  # Nome do funcionário (última parte)

                    if unidade == "minutos":
                        if len(partes) < 5:
                            raise ValueError("Formato inválido para callback_data de minutos.")
                        tempo = f"{partes[2]} a {partes[3]}"  # Minutos formatados corretamente
                    elif unidade == "horas":
                        if len(partes) < 4:
                            raise ValueError("Formato inválido para callback_data de horas.")
                        tempo = f"{partes[2]}"  # Horas formatadas corretamente
                    else:
                        raise ValueError("Unidade de tempo inválida.")

                    # Verificar e inicializar o dicionário de ausências para o usuário
                    if usuario_id not in usuarios:
                        usuarios[usuario_id] = {}
                    if "ausencias" not in usuarios[usuario_id]:
                        usuarios[usuario_id]["ausencias"] = {}

                    if funcionario not in usuarios[usuario_id]["ausencias"]:
                        usuarios[usuario_id]["ausencias"][funcionario] = {}

                    # Registrar o tempo de atraso
                    if "tempo_atraso" not in usuarios[usuario_id]["ausencias"][funcionario]:
                        usuarios[usuario_id]["ausencias"][funcionario]["tempo_atraso"] = {}

                    if unidade == "minutos":
                        usuarios[usuario_id]["ausencias"][funcionario]["tempo_atraso"]["tempo_atraso_minutos"] = f"{tempo} minutos"
                    else:
                        usuarios[usuario_id]["ausencias"][funcionario]["tempo_atraso"]["tempo_atraso_horas"] = f"{tempo} horas"

                    print(f"Tempo de atraso registrado: {tempo} para {funcionario}")

                    # Perguntar o motivo do atraso
                    markup = InlineKeyboardMarkup(row_width=1)
                    markup.add(
                        InlineKeyboardButton("Transporte", callback_data=f"atrasomot_transporte_{funcionario}"),
                        InlineKeyboardButton("Trânsito", callback_data=f"atrasomot_transito_{funcionario}"),
                        InlineKeyboardButton("Assunto Familiar", callback_data=f"atrasomot_familiar_{funcionario}")
                    )


                    bot.send_message(
                        usuario_id,
                        f"O tempo de atraso de {funcionario} foi de {tempo}. Qual o motivo do atraso?",
                        reply_markup=markup
                    )

                except Exception as e:
                    print(f"Erro ao salvar o tempo de atraso: {e}")
                    bot.send_message(usuario_id, f"Ocorreu um erro ao salvar o tempo de atraso: {e}")

            # Registrar motivo do atraso selecionado pelo usuário
            @bot.callback_query_handler(func=lambda call: call.data.startswith("atrasomot_"))
            def registrar_motivo_atraso(call):
                print(f"Callback recebido: {call.data}")  # Verificando a callback_data
                usuario_id = call.message.chat.id
                try:
                    # Dividir o callback_data corretamente
                    partes = call.data.split("_")
                    print(f"Partes divididas: {partes}")  # Para ver se a divisão está funcionando corretamente
                    if len(partes) < 3:
                        raise ValueError("Formato inválido para callback_data.")

                    motivo = partes[1]  # Motivo do atraso
                    funcionario = "_".join(partes[2:])  # Nome do funcionário (pode conter underscores)

                    # Verificar se o dicionário de ausências está inicializado
                    if "ausencias" not in usuarios[usuario_id]:
                        usuarios[usuario_id]["ausencias"] = {}

                    if funcionario not in usuarios[usuario_id]["ausencias"]:
                        usuarios[usuario_id]["ausencias"][funcionario] = {}

                    # Registrar o motivo do atraso
                    usuarios[usuario_id]["ausencias"][funcionario]["motivo_atraso"] = motivo.capitalize()
                     # Log para depuração
                    print(f"Registrando motivo de atraso: {motivo.capitalize()} para o funcionário {funcionario}")

                    # Enviar mensagem de confirmação
                    bot.send_message(
                        usuario_id,
                        f"Atraso de {funcionario} registrado com motivo: {motivo.capitalize()}.",
                        reply_markup=menu_principal()  # Certifique-se de que o menu_principal() está retornando um InlineKeyboardMarkup
                    )

                except Exception as e:
                    bot.send_message(usuario_id, f"Ocorreu um erro ao registrar o motivo do atraso: {e}")

            #Função para tratar a opção atestado
            @bot.callback_query_handler(func=lambda call: call.data.startswith("atestado_"))
            def tratar_atestado(call):
                _,opcao,funcionario= call.data.split("_")
                usuario_id = call.message.chat.id

                if opcao == "agora":
                    bot.send_message(
                        usuario_id,
                        f"Por favor, envie o atestado médico de {funcionario} como documento ou imagem."
                    )
                    bot.register_next_step_handler_by_chat_id(
                        usuario_id, receber_atestado, funcionario
                    )
                elif opcao == "depois":
                    usuarios[usuario_id]["ausencias"][funcionario]["prazo_atestado"] = datetime.now().timestamp() + 86400
                    bot.send_message(
                        usuario_id,
                        "Você tem 48 horas para enviar o atestado. Caso contrário, será contado como falta.",
                        reply_markup=menu_principal()
                    )
                elif opcao =="ja":
                    bot.send_message(
                        usuario_id,
                        f"Atestado já foi enviado previamente.",
                        reply_markup=menu_principal()
                    )

            #função para receber o atestado
            def receber_atestado(message, funcionario):
                usuario_id = message.chat.id

                # Armazenar message e funcionario no dicionário de usuários
                if usuario_id not in usuarios:
                    usuarios[usuario_id] = {"ausencias": {}}
                usuarios[usuario_id]["message"] = message
                usuarios[usuario_id]["funcionario"] = funcionario

                #verificar se o arquivo foi enviado
                if message.content_type in ["document", "photo"]:

                    # Registrar o atestado no banco de dados (estrutura de dados dos usuários)
                    usuarios[usuario_id]["ausencias"][funcionario]["atestado"] = True
                    bot.send_message(
                        usuario_id,
                        f"Atestado de {funcionario} recebido com sucesso.",
                        reply_markup=menu_principal()
                    )

                    
                else:
                    bot.send_message(
                        usuario_id,
                        f"Por favor, envie o atestado como documento ou imagem."
                    )
                    bot.register_next_step_handler_by_chat_id(
                        usuario_id, receber_atestado, funcionario
                    )

            #handler para processar se ausênciafoi acordada (Sim/Não)
            @bot.callback_query_handler(func=lambda call: call.data.startswith("ausente_"))
            def ausencia_acordada(call):
                _, resposta, funcionario = call.data.split("_")
                usuario_id = call.message.chat.id

                # Registrar a resposta da falta acordada
                usuarios[usuario_id]["ausencias"][funcionario]["acordado"] = resposta == "sim"

                # Mensagem de resposta
                if resposta == "sim":
                    bot.send_message(
                        usuario_id,
                        f"A falta do {funcionario} foi acordada, por favor, justifique essa ausência.",
                        reply_markup=InlineKeyboardMarkup(row_width=2)  # 2 colunas
                        .add(
                            InlineKeyboardButton("Desligado da empresa", callback_data=f"justificativa_desligado_{funcionario}"),
                            InlineKeyboardButton("Assunto pessoal", callback_data=f"justificativa_assunto_pessoal_{funcionario}"),
                            InlineKeyboardButton("Não teve justificativa", callback_data=f"justificativa_sem_justificativa_{funcionario}"),
                            InlineKeyboardButton("Consulta médica", callback_data=f"justificativa_consulta_medica_{funcionario}"),
                            InlineKeyboardButton("Outro (falar com o RH)", callback_data=f"justificativa_outra_justificativa_{funcionario}")
                        )
                    )
                else:
                    bot.send_message(
                        usuario_id,
                        f"A falta do {funcionario} não foi acordada, por favor, justifique essa ausência.",
                        reply_markup=InlineKeyboardMarkup(row_width=2)  # 2 colunas
                        .add(
                            InlineKeyboardButton("Desligado da empresa", callback_data=f"justificativa_desligado_empresa_{funcionario}"),
                            InlineKeyboardButton("Assunto pessoal", callback_data=f"justificativa_assunto_pessoal_{funcionario}"),
                            InlineKeyboardButton("Não teve justificativa", callback_data=f"justificativa_sem_justificativa_{funcionario}"),
                            InlineKeyboardButton("Consulta médica", callback_data=f"justificativa_consulta_medica_{funcionario}"),
                            InlineKeyboardButton("Outro (falar com o RH)", callback_data=f"justificativa_outra_justificativa_{funcionario}")
                        )
                    )

            @bot.callback_query_handler(func=lambda call: call.data.startswith("justificativa_"))
            def registrar_justificativa(call):
                usuario_id = call.message.chat.id
                partes = call.data.split("_")

                # Verificar se o número de partes é suficiente para processar
                if len(partes) < 4:
                    bot.send_message(usuario_id, "Erro: Formato de callback_data inválido.")
                    return

                # Posição 1 e 2 - Motivo escolhido
                motivo = " ".join(partes[1:3])  # Motivo é formado pela junção de 1 e 2 (ex: 'desligado da empresa')

                # Posição 3 é o nome do funcionário
                funcionario = partes[3]  # Nome do funcionário (sem underscores, já que usamos para separar os dados)

                # Registrar a justificativa no dicionário de ausências
                if "ausencias" not in usuarios[usuario_id]:
                    usuarios[usuario_id]["ausencias"] = {}

                if funcionario not in usuarios[usuario_id]["ausencias"]:
                    usuarios[usuario_id]["ausencias"][funcionario] = {}

                usuarios[usuario_id]["ausencias"][funcionario]["justificativa"] = motivo

                # Enviar mensagem de confirmação
                bot.send_message(
                    usuario_id,
                    f"A justificativa de {funcionario} foi registrada como: {motivo.replace('_', ' ')}.",
                    reply_markup=menu_principal()  # Retorna ao menu principal após registrar a justificativa
                )



            # Função para registrar tempo de "Saiu mais cedo"
            @bot.callback_query_handler(func=lambda call: call.data.startswith("cedo_"))
            def registrar_tempo_saiu_mais_cedo(call):
                usuario_id = call.message.chat.id
                try:
                    # Dividir o callback_data
                    _, tipo, funcionario = call.data.split("_")
                    
                    # Verificar se o "funcionario" foi extraído corretamente
                    if not funcionario:
                        raise ValueError("Funcionário não especificado.")
                    
                    if tipo == "horas":
                        markup = InlineKeyboardMarkup(row_width=1)
                        markup.add(
                            InlineKeyboardButton("1 hora", callback_data=f"cedotemp_horas_1_{funcionario}"),
                            InlineKeyboardButton("2 horas", callback_data=f"cedotemp_horas_2_{funcionario}"),
                            InlineKeyboardButton("3 horas ou mais", callback_data=f"cedotemp_horas_3_{funcionario}")
                        )
                        bot.send_message(
                            usuario_id,
                            f"Selecione o tempo que {funcionario} saiu mais cedo:",
                            reply_markup=markup
                        )

                    elif tipo == "minutos":
                        markup = InlineKeyboardMarkup(row_width=1)
                        markup.add(
                            InlineKeyboardButton("1 a 15 minutos", callback_data=f"cedotemp_minutos_1_15_{funcionario}"),
                            InlineKeyboardButton("15 a 30 minutos", callback_data=f"cedotemp_minutos_15_45_{funcionario}"),
                            InlineKeyboardButton("30 a 59 minutos", callback_data=f"cedotemp_minutos_30_59_{funcionario}")
                        )
                        bot.send_message(
                            usuario_id,
                            f"Selecione o tempo que {funcionario} saiu mais cedo:",
                            reply_markup=markup
                        )
                
                except ValueError as e:
                    bot.send_message(usuario_id, f"Erro: {e}")
                    return

            @bot.callback_query_handler(func=lambda call: call.data.startswith("cedotemp_"))
            def salvar_tempo_mais_cedo(call):
                usuario_id = call.message.chat.id
                try:
                    # Dividir o callback_data corretamente
                    partes = call.data.split("_")
                    print(f"Partes divididas: {partes}")  # Para depuração

                    if len(partes) < 4:
                        raise ValueError("Formato inválido para callback_data.")

                    unidade = partes[1]  # Unidade de tempo (horas ou minutos)
                    tempo = None
                    funcionario = partes[-1]  # Nome do funcionário (última parte)

                    if unidade == "minutos":
                        if len(partes) < 5:
                            raise ValueError("Formato inválido para callback_data de minutos.")
                        tempo = f"{partes[2]} a {partes[3]}"  # Minutos formatados corretamente
                    elif unidade == "horas":
                        if len(partes) < 4:
                            raise ValueError("Formato inválido para callback_data de horas.")
                        tempo = f"{partes[2]}"  # Horas formatadas corretamente
                    else:
                        raise ValueError("Unidade de tempo inválida.")

                    # Verificar se o dicionário de ausências está inicializado
                    if usuario_id not in usuarios:
                        usuarios[usuario_id] = {}
                    if "ausencias" not in usuarios[usuario_id]:
                        usuarios[usuario_id]["ausencias"] = {}

                    if funcionario not in usuarios[usuario_id]["ausencias"]:
                        usuarios[usuario_id]["ausencias"][funcionario] = {}

                    # Registrar o tempo que o funcionário saiu mais cedo
                    if "tempo_saiu_mais_cedo" not in usuarios[usuario_id]["ausencias"][funcionario]:
                        usuarios[usuario_id]["ausencias"][funcionario]["tempo_saiu_mais_cedo"] = {}

                    if unidade == "minutos":
                        usuarios[usuario_id]["ausencias"][funcionario]["tempo_saiu_mais_cedo"]["tempo_saiu_minutos"] = f"{tempo} minutos"
                    else:
                        usuarios[usuario_id]["ausencias"][funcionario]["tempo_saiu_mais_cedo"]["tempo_saiu_horas"] = f"{tempo} horas"

                    print(f"Tempo de saída mais cedo registrado: {tempo} para {funcionario}")

                    # Perguntar o motivo de sair mais cedo
                    markup = InlineKeyboardMarkup(row_width=1)
                    markup.add(
                        InlineKeyboardButton("Assunto Familiar", callback_data=f"cedomot_familiar_{funcionario}"),
                        InlineKeyboardButton("Urgência", callback_data=f"cedomot_urgencia_{funcionario}"),
                        InlineKeyboardButton("Combinou com o gerente", callback_data=f"cedomot_combinado_{funcionario}")
                    )
                    

                    bot.send_message(
                        usuario_id,
                        f"O tempo que {funcionario} saiu mais cedo foi de {tempo} {unidade}. Qual o motivo?",
                        reply_markup=markup
                    )

                except Exception as e:
                    print(f"Erro ao salvar o tempo de saída: {e}")
                    bot.send_message(usuario_id, f"Ocorreu um erro ao salvar o tempo de saída: {e}")

            # Função para registrar a justificativa de "Saiu mais cedo"
            @bot.callback_query_handler(func=lambda call: call.data.startswith("cedomot_"))
            def registrar_justificativa_saiu_mais_cedo(call):
                usuario_id = call.message.chat.id
                try:
                    # Dividir o callback_data corretamente
                    partes = call.data.split("_")
                    if len(partes) < 3:
                        raise ValueError("Formato inválido para callback_data.")

                    cedomotivo = partes[1]  # Motivo de sair mais cedo
                    funcionario = "_".join(partes[2:])  # Nome do funcionário (pode conter underscores)

                    # Verificar se o dicionário de ausências está inicializado
                    if "ausencias" not in usuarios[usuario_id]:
                        usuarios[usuario_id]["ausencias"] = {}

                    if funcionario not in usuarios[usuario_id]["ausencias"]:
                        usuarios[usuario_id]["ausencias"][funcionario] = {}

                    # Registrar a justificativa de sair mais cedo
                    usuarios[usuario_id]["ausencias"][funcionario]["justificativa_cedo"] = cedomotivo.capitalize()

                    # Enviar mensagem de confirmação
                    bot.send_message(
                        usuario_id,
                        f"A saída antecipada de {funcionario} foi registrada com o motivo: {cedomotivo.capitalize()}.",
                        reply_markup=menu_principal()
                    )
                except Exception as e:
                    bot.send_message(usuario_id, f"Ocorreu um erro ao registrar a justificativa: {e}")
                
        

            # Função para excluir arquivos antigos, com verificação para não excluir arquivos essenciais
            def excluir_arquivos_antigos(diretorio):
                """ Exclui arquivos do diretório especificado, exceto os protegidos, apenas no horário definido. """
                
                # Lista de arquivos que NÃO devem ser excluídos
                arquivos_preservados = {"bot.py", ".env", "text.txt"}

                # Obtém a hora atual
                agora = datetime.now()

                # Verifica se é o horário definido para exclusão (exemplo: 14:05)
                if agora.hour == 13 and agora.minute == 5:
                    for nome_arquivo in os.listdir(diretorio):
                        caminho_arquivo = os.path.join(diretorio, nome_arquivo)

                        # Se o arquivo não estiver na lista de preservados, exclui
                        if nome_arquivo not in arquivos_preservados and os.path.isfile(caminho_arquivo):
                            try:
                                os.remove(caminho_arquivo)
                                print(f"Arquivo {nome_arquivo} excluído com sucesso.")
                            except Exception as e:
                                print(f"Erro ao excluir {nome_arquivo}: {e}")

            def monitorar_exclusao(diretorios):
                """ Verifica continuamente a hora e exclui arquivos apenas no horário programado. """
                while True: 
                    agora = datetime.now()
                    
                    # Só roda a exclusão no horário corret
                    if agora.hour == 13 and agora.minute == 5:
                        for diretorio in diretorios:
                            if os.path.exists(diretorio):
                                excluir_arquivos_antigos(diretorio)
                    
                    time.sleep(60)  # Aguarda 60 segundos antes de verificar novamente
                    
            # Criar thread separada para monitoramento
            diretorio_relatorios = ["./","./arquivos_individuais"]  # Caminho correto dos arquivos
            thread_exclusao = threading.Thread(target=monitorar_exclusao, args=(diretorio_relatorios,))
            thread_exclusao.daemon = True
            thread_exclusao.start()

            # Handler para finalizar a conversa
            @bot.callback_query_handler(func=lambda call: call.data == "finalizar_conversa")
            def finalizar_conversa(call):
                usuario_id = call.message.chat.id
                nome_loja = lojasCadastradas[usuarios[usuario_id]['loja']]['nome']
                # Obter message e funcionario armazenados anteriormente
                message = usuarios[usuario_id].get("message")
                funcionario = usuarios[usuario_id].get("funcionario")

                # Chamar as funções para gerar os relatórios e salvar na pasta correta
                #gerar_relatorio_pdf(usuario_id)
                gerar_relatorio_excel(usuario_id)

                # Verificar se há informações para armazenar
                if "ausencias" in usuarios[usuario_id] and usuarios[usuario_id]["ausencias"]:
                    loja = usuarios[usuario_id]["loja"]
                    
                    # Inicializar a lista de ausências para a loja, se necessário
                    if loja not in informacoes_diarias:
                        informacoes_diarias[loja] = []
                    
                    # Adicionar as ausências atuais ao dicionário global
                    informacoes_diarias[loja].append(usuarios[usuario_id]["ausencias"])

                # Enviar mensagem finalizando a conversa
                bot.send_message(usuario_id, "Relatório gerado. O relatório foi enviado ao responsável. Mande qualquer mensagem para iniciar uma nova conversa.")
                
                responsavel_ids = [6566910217]
                for responsavel_id in responsavel_ids:
                    # Gerar o conteúdo do relatório
                    relatorio_conteudo = extrair_conteudo_relatorio(usuario_id)               
    
                    # Enviar mensagem finalizando a conversa com o relatório
                    bot.send_message(responsavel_id, f"Relatório gerado {nome_loja} de {usuario_id}:\n\n{relatorio_conteudo}")

                # Enviar o atestado, se houver
                    if "atestado" in usuarios[usuario_id]["ausencias"].get(funcionario, {}):
                        if message.content_type == "document":
                            try:
                                bot.send_document(responsavel_id, message.document.file_id)
                                bot.send_message(responsavel_id, f"Atestado de {funcionario} enviado.")
                            except Exception as e:
                                bot.send_message(responsavel_id, f"Erro ao enviar o atestado como documento: {e}")
                        elif message.content_type == "photo":
                            bot.send_photo(responsavel_id, message.photo[-1].file_id)
                            bot.send_message(responsavel_id, f"Atestado de {funcionario} enviado.")

            # Função para extrair o conteúdo do relatório gerado para a mensagem
            def extrair_conteudo_relatorio(usuario_id):
                # Obter nome da loja
                loja = lojasCadastradas[usuarios[usuario_id]['loja']]['nome']
                funcionarios_ausentes = usuarios[usuario_id]["ausencias"]  # Exemplo de como extrair informações

                # Criando o conteúdo do relatório como uma string
                conteudo = f"Relatório da Loja: {loja}\n\n"
                if not usuarios[usuario_id]["ausencias"]: # Se não houver funcionários ausentes, exibe "Todos entraram no horário normal."
                    conteudo += f"Todos entraram no horário normal."
                # Iterar sobre os dados dos funcionários
                for funcionario, dados in funcionarios_ausentes.items():
                    conteudo += f"Funcionário: {funcionario}\n"
                    
                    # Adicionar motivo de ausência, se presente
                    if 'motivo' in dados:
                        conteudo += f"Motivo: {dados['motivo']}\n"

                    # Adicionar tempo de atraso, se presente
                    if 'tempo_atraso_horas' in dados.get('tempo_atraso', {}):
                        conteudo += f"Tempo de Atraso: {dados['tempo_atraso']['tempo_atraso_horas']} horas\n"
                    if 'tempo_atraso_minutos' in dados.get('tempo_atraso', {}):
                        conteudo += f"Tempo de Atraso: {dados['tempo_atraso']['tempo_atraso_minutos']} minutos\n"
                    
                    # Adicionar justificativa de atraso, se presente
                    if 'motivo_atraso' in dados:
                        conteudo += f"Justificativa de Atraso: {dados['motivo_atraso']}\n"

                    # Adicionar tempo de saída mais cedo, se presente
                    if 'tempo_saiu_horas' in dados.get('tempo_saiu_mais_cedo', {}):
                        conteudo += f"Tempo de Saída Mais Cedo: {dados['tempo_saiu_mais_cedo']['tempo_saiu_horas']} horas\n"
                    if 'tempo_saiu_minutos' in dados.get('tempo_saiu_mais_cedo', {}):
                        conteudo += f"Tempo de Saída Mais Cedo: {dados['tempo_saiu_mais_cedo']['tempo_saiu_minutos']} minutos\n"
                    
                    # Adicionar justificativa por sair mais cedo, se presente
                    if 'justificativa_cedo' in dados:
                        conteudo += f"Justificativa por sair mais cedo: {dados['justificativa_cedo']}\n"

                    # Verificar e adicionar informações sobre atestado
                    if 'atestado' in dados:
                        conteudo += "Atestado: Enviado\n"
                    if 'prazo_atestado' in dados:
                        conteudo += "Atestado: Deve ser enviado em 48 horas\n"

                    # Indicar se a ausência foi acordada
                    if 'acordado' in dados:
                        acordado = "Sim" if dados["acordado"] else "Não"
                        conteudo += f"Falta Acordada: {acordado}\n"

                    # Adicionar justificativa, se presente
                    if 'justificativa' in dados:
                        conteudo += f"Justificativa: {dados['justificativa']}\n"
                    
                    # Linha divisória entre os funcionários
                    conteudo += "---------\n"
                
                return conteudo

            def gerar_e_enviar_relatorios_consolidados():
                print("Iniciando geração e envio de relatórios consolidados...")

                # Consolidar os relatórios
                gerar_relatorio_excel_consolidado()
                #gerar_relatorio_pdf_consolidado()

                # IDs dos responsáveis para envio
                responsavel_ids = [6566910217]

                for responsavel_id in responsavel_ids:
                    # Enviar Excel consolidado
                    try:
                        with open("relatorio_consolidado_todas_as_lojas.xlsx", "rb") as excel_file:
                            bot.send_document(responsavel_id, excel_file)
                    except Exception as e:
                        bot.send_message(responsavel_id, f"Erro ao enviar Excel consolidado: {e}")

                print("Relatórios consolidados enviados com sucesso.")


            def enviar_relatorios_consolidados_para_responsavel(nome_relatorio, df_final):
                responsavel_ids = [6566910217]  # IDs dos responsáveis
                for responsavel_id in responsavel_ids:

                    # Enviar o Excel consolidado
                    try:
                        with open("relatorio_consolidado_todas_as_lojas.xlsx", "rb") as excel_file:
                            bot.send_document(responsavel_id, excel_file)
                    except Exception as e:
                        bot.send_message(responsavel_id, f"Erro ao enviar Excel consolidado: {e}")

            def gerar_relatorio_excel_consolidado():
                arquivos_xlsx = [f for f in os.listdir(PASTA_ARQUIVOS) if f.endswith(".xlsx")]
                if not arquivos_xlsx:
                    print("Nenhum arquivo Excel encontrado para consolidar.")
                    return

                dataframes = []
                for arquivo in arquivos_xlsx:
                    caminho = os.path.join(PASTA_ARQUIVOS, arquivo)
                    try:
                        df = pd.read_excel(caminho)
                        dataframes.append(df)
                    except Exception as e:
                        print(f"Erro ao carregar {arquivo}: {e}")

                if not dataframes:
                    print("Nenhum dado carregado para consolidar.")
                    return

                # Concatenando todos os DataFrames em um só
                df_final = pd.concat(dataframes, ignore_index=True)

                # Nome do arquivo consolidado
                nome_arquivo_consolidado = "relatorio_consolidado_todas_as_lojas.xlsx"
                df_final.to_excel(nome_arquivo_consolidado, index=False)

                print(f"Relatório Excel consolidado salvo como {nome_arquivo_consolidado}.")
                ajustar_largura_colunas(nome_arquivo_consolidado)

            # Agendar a execução da função ao final do dia (por exemplo, às 23:59)
            schedule.every().day.at("13:03").do(gerar_e_enviar_relatorios_consolidados)
            # Resetar o dicionário no início do dia
            schedule.every().day.at("13:03").do(lambda: informacoes_diarias.clear())


            # Função para verificar relatórios faltantes
            def verificar_relatorios_faltantes():
                # Array com os números das lojas que devem enviar relatórios
                lojas_esperadas = ["1", "4", "5","6","7"]  # Substitua pelos números reais das lojas

                # IDs dos responsáveis que devem ser notificados
                responsavel_ids = [6566910217,1243989891]  # Substitua pelos IDs reais dos responsáveis

                # Lista para armazenar as lojas que enviaram relatórios
                lojas_enviadas = []

                # Verificar os arquivos na pasta "arquivos_individuais"
                for nome_arquivo in os.listdir(PASTA_ARQUIVOS):
                    if nome_arquivo.endswith(".xlsx"):
                        # Extrair o número da loja do nome do arquivo
                        partes = nome_arquivo.split("_")
                        if len(partes) >= 2:  # Verifica se o nome do arquivo está no formato esperado
                            numero_loja = partes[1].replace("Loja", "")  # Remove "Loja" para obter apenas o número
                            if numero_loja in lojas_esperadas:
                                lojas_enviadas.append(numero_loja)

                # Verificar quais lojas não enviaram relatórios
                lojas_faltantes = [loja for loja in lojas_esperadas if loja not in lojas_enviadas]

                # Se houver lojas faltantes, enviar uma mensagem para os responsáveis
                if lojas_faltantes:
                    mensagem = "⚠️ Atenção!, As seguintes lojas não enviaram relatórios:\n\n"
                    mensagem += "\n".join([f"Loja {loja}" for loja in lojas_faltantes])
                    mensagem += "\n\nPor favor, verifiquem."

                    for responsavel_id in responsavel_ids:
                        try:
                            bot.send_message(responsavel_id, mensagem)
                        except Exception as e:
                            print(f"Erro ao enviar mensagem para {responsavel_id}: {e}")
                else:
                    print("Todas as lojas enviaram relatórios.")

            # Agendar a verificação de relatórios faltantes
            schedule.every().day.at("13:04").do(verificar_relatorios_faltantes)

            #Reiniciar a conversa a qualquer momento
            @bot.message_handler(func=lambda mensagem: usuarios.get(mensagem.chat.id, {}).get("finalizado", False))
            def reiniciar_conversa(mensagem):
                usuario_id = mensagem.chat.id

                # Apagar completamente os dados do usuário
                if usuario_id in usuarios:
                    del usuarios[usuario_id]  # Remove os dados do usuário completamente
                # Reiniciar a conversa chamando iniciar_conversa novamente
                iniciar_conversa(mensagem)

            # Comando /reiniciar chama a função reiniciar_conversa
            @bot.message_handler(commands=['reiniciar'])
            def retomar_conversa(message):
                reiniciar_conversa(message)

            bot.polling()
        except Exception as e:
            print(f"Ocorreu um erro: {e}")
            print("Reiniciando a aplicação em 5 segundos...")
            time.sleep(5)
            os.execv(sys.executable, [sys.executable] + sys.argv)  # Reinicia o script

if __name__ == "__main__":
    main()